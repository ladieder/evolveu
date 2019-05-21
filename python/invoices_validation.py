from openpyxl import load_workbook

def invoices_validation(workbook):
    wb = load_workbook(workbook)

    print("data validation for ", workbook)

    sheet_names = wb.sheetnames
    cust_ws = wb["customers"]
    inv_ws = wb["invoices"]
    lines_ws = wb["line_items"]
    prod_ws = wb["products"]

    # check 'Customers' sheet for duplicate customer ids and for blank customer ids or customer names
    cust_dict = {}
    for row in cust_ws.iter_rows(min_row=2, max_col=2, values_only=True):
        # if cust_id and cust_name are empty, skip blank row (do nothing)
        if row[0] == None and row[1] == None:
            pass

        # else if cust_id is empty, don't add customer to cust_dict and give error message
        elif row[0] == None:
            print("customer name ", row[1], " has no customer id. Customer not added.")

        # else if check if the cust_id is an integer, if it's not integer, don't add and give error message
        elif type(row[0]) != int:
            print("customer id ", row[0], " is not an integer. Customer not added.")

        # else (cust_id is not empty, cust_name may or may not be blank)
        else:
            # check for duplicate customer ids
            # if cust_id is already key in cust_dict, don't add to cust_dict and give error message (don't bother checking if cust_name is blank or not)
            if row[0] in cust_dict:
                print("duplicate customer with id of: ", row[0], ". Duplicate customer not added.")

            # if cust_id is not duplicate, check if cust_name is blank
            # else if cust_name is empty, add customer with no name and give error message
            elif row[1] == None:
                cust_dict[row[0]] = row[1]
                print("customer with id of ", row[0], " has no name specified. Customer id added with no name.") 

            # else (cust_id is not duplicated and cust_name is not blank), add customter to cust_dict
            else:
                cust_dict[row[0]] = row[1]

    print(cust_dict)

    # check 'Products'sheet for duplicate product ids
    prod_dict = {}
    for row in prod_ws.iter_rows(min_row=2, max_col=4, values_only=True):
        # skip blank rows
        if row[0] == None and row[1] == None and row[2] == None:
            pass

        # else if prod_id is duplicate, don't add to prod_dict and give error message
        elif row[0] in prod_dict:
            print("duplicate product with id of: ", row[0], ". Duplicate prod_id not added.")

        # else (prod_id is not duplicate), add to prod_dict
        else:
            prod_dict[row[0]] = {"prod_name": row[1], "prod_cost": row[2]}

    # check 'Invoices' sheet for duplicate invoices
    inv_dict = {}
    for row in inv_ws.iter_rows(min_row=2, max_col=3, values_only=True):
        # skip blank rows
        if row[0] == None and row[1] == None and row[2] == None:
            pass
        
        # else if inv_id is duplicate, don't add and give error message
        elif row[0] in inv_dict:
            print("duplicate invoice with id of: ", row[0], ". Duplicate inv_id not added.")

        # else (inv_id is not duplicate), add inv_id, cust_id and inv_total to inv_dict
        else:
            inv_dict[row[0]] = {"cust_id": row[1], "invoice_date": row[2]}

    # add line items (prod_id, quantity and sub_total) and inv_total to inv_dict for each invoice
    inv_total = 0
    for row in lines_ws.iter_rows(min_row=2, max_col=3, values_only=True):
        # skip blank rows
        if row[0] == None and row[1] == None and row[2] == None:
            pass
        
        # else if a line item already exists in inv_dict for specified inv_id, set temp line_list to whatever already exists in inv_dict
        elif "line_items" in inv_dict[row[0]]:
            line_list = inv_dict[row[0]]["line_items"]

        # else (specified inv_id does not already have a line item in inv_dict), reset temp line_list to blank and inv_total to 0
        else:
            line_list = []
            inv_total = 0
        
        # calculate sub_total for line item, increment invoice sub_total by new line item
        sub_total = row[2] * prod_dict[row[1]]["prod_cost"]
        inv_total += sub_total

        # append new line item to line_list, overwrite line_items in inv_dict with temp line_list, overwrite inv_total with current value
        line_list.append({"prod_id": row[1], "quantity": row[2], "sub_total": sub_total})
        inv_dict[row[0]]["line_items"] = line_list
        inv_dict[row[0]]["inv_total"] = inv_total

    # check number of invoices and print result
    num_invoices = len(inv_dict)
    print("number of invoices: ", num_invoices)

    # check total of all invoices
    month_total = 0
    for item in inv_dict:
        month_total += inv_dict[item]["inv_total"]
    print("total of all invoices: ", month_total)

    data_dict = {"cust_dict": cust_dict, "prod_dict": prod_dict, "inv_dict": inv_dict, "sheet_names": sheet_names}

    return data_dict