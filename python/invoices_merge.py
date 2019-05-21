import pprint
from openpyxl import load_workbook
from invoices_validation import invoices_validation

def invoices_merge(wb_merge):
    wb_master = 'invoices_master.xlsx'
    dict_merge = invoices_validation(wb_merge)
    dict_master = invoices_validation(wb_master)

    print("dict_merge")
    pprint.pprint(dict_merge)
    print("dict_master")
    pprint.pprint(dict_master)

    # dict_merged = {}
    if dict_master["sheet_names"] != dict_merge["sheet_names"]:
        print("error - workbook sheet names do not match")
    else:
        dict_master["cust_dict"].update(dict_merge["cust_dict"])
        print("dict_master updated")
        pprint.pprint(dict_master)

        # print updated dict to master workbook
        wb = load_workbook(wb_master)
        cust_ws = wb["customers"]
        inv_ws = wb["invoices"]
        lines_ws = wb["line_items"]
        prod_ws = wb["products"]

        for i, item in enumerate(dict_master["cust_dict"]):
            cust_ws.cell(row=i+2, column=1, value=item)
            cust_ws.cell(row=i+2, column=2, value=dict_master["cust_dict"][item])

        for i, item in enumerate(dict_master["inv_dict"]):
            inv_ws.cell(row=i+2, column=1, value=item)
            inv_ws.cell(row=i+2, column=2, value=dict_master["inv_dict"][item]["cust_id"])
            inv_ws.cell(row=i+2, column=3, value=dict_master["inv_dict"][item]["invoice_date"])

        for i, item in enumerate(dict_master["inv_dict"]):
            for line in dict_master["inv_dict"][item]["line_items"]:
                lines_ws.cell(row=i+2, column=1, value=item)
                lines_ws.cell(row=i+2, column=2, value=line["prod_id"])
                lines_ws.cell(row=i+2, column=3, value=line["quantity"])
                i = i + 1

        for i, item in enumerate(dict_master["prod_dict"]):
            prod_ws.cell(row=i+2, column=1, value=item)
            prod_ws.cell(row=i+2, column=2, value=dict_master["prod_dict"][item]["prod_name"])
            prod_ws.cell(row=i+2, column=3, value=dict_master["prod_dict"][item]["prod_cost"])

        wb.save(wb_master)

invoices_merge('invoices_feb.xlsx')