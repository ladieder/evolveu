from openpyxl import load_workbook
wbmerge = load_workbook('invoices.xlsx')
wbmaster = load_workbook('invoices_master.xlsx')








l = [[1, 2, 3], [2, 4, 5], [1, 2, 3], [2, 4, 5]]
# s = set(tuple(i) for i in l)
s = set()
for i in l:
    s.add(tuple(i))

print(s)